from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl
from bs4 import BeautifulSoup
import time


# Create an Excel workbook and set the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set the header for the "Attacking Rackets" column
sheet["A1"] = "Attacking Rackets"
sheet["B1"] = "Price($)"
all_text = []
all_price = []


# Update to the correct path
# Path to the Chrome WebDriver


# Set up the Chrome WebDriver instance in headless mode
options = Options()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
driver = webdriver.Chrome()

# Navigate to the website
driver.get("https://in.lining.studio/category/badminton/rackets/attacking")

# Wait for the page to load (you may need to customize the wait time)
driver.implicitly_wait(10)

# Function to scroll down the page


def scroll_down():
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)  # Adjust sleep time as needed


number = 1
while number <= 12:
    driver.get(
        f"https://in.lining.studio/category/badminton/rackets/attacking?filterBy=&sortBy=typesense.isOOS%3Aasc&page={number}&searchQuery=%2Fbadminton%2Frackets%2Fattacking%2F&sidebar=false")

    # Get the initial page source after interacting with the website
    page_source = driver.page_source

    # Use BeautifulSoup to parse the HTML
    soup = BeautifulSoup(page_source, "html.parser")

    # Find all elements with the specified CSS selector
    nextpage_elements = driver.find_elements(
        By.CLASS_NAME, "text-sm.font-semibold.leading-snug.text-black.md\\:text-base")

    price_element = driver.find_elements(
        By.CLASS_NAME, "card-price")

    # Write the text content of each element to the Excel sheet
    for index, element in enumerate(nextpage_elements, start=2):
        all_text.append(element.text)

    # Write the text content of each price element to the Excel sheet
    for index, element in enumerate(price_element, start=2):
        all_price.append(element.text)
    # Scroll down and load more content
    scroll_down()

    # ELEMENTS AFTER SCROLL:

    # Find elements again after scrolling
    elements_after_scroll = driver.find_elements(
        By.CLASS_NAME, "text-sm.font-semibold.leading-snug.text-black.md\\:text-base")

    # find price elements after scrolling
    price_elements_after_scroll = driver.find_elements(
        By.CLASS_NAME, "card-price")

    # APPENDING THEM TO THE ARRAYS:

   # Write the text content of the additional elements to the Excel sheet
    for index, element in enumerate(elements_after_scroll, start=len(nextpage_elements) + 2):
        all_text.append(element.text if element.text else "out of stock")

   # Write the text content of the additional elements to the Excel sheet
for index, element in enumerate(price_elements_after_scroll, start=len(price_element) + 2):
        all_price.append(element.text if element.text else "out of stock")

    number += 1


# Write the array to the Excel sheet
for index, text_content in enumerate(all_text, start=2):
    sheet.cell(row=index, column=1, value=text_content)

for index, price_content in enumerate(all_price, start=2):
    sheet.cell(row=index, column=2, value=price_content)


workbook.save("scrape.xlsx")

# Close the WebDriver
driver.quit()
