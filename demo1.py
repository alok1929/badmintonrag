from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import openpyxl
from bs4 import BeautifulSoup
import time

# Create an Excel workbook and set the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set the header for the "Defensive Rackets" column
sheet["A1"] = "Defensive Rackets"
sheet["B1"] = "Price($)"
all_text = []
all_price = []

# Set up Chrome options and driver
options = Options()
options.add_argument("--no-sandbox")
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

# Navigate to the website
driver.get("https://in.lining.studio/category/badminton/rackets/defensive")

# Wait for the page to load
driver.implicitly_wait(10)


def scroll_down():
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)


number = 1
while number <= 4:
    driver.get(
        f"https://in.lining.studio/category/badminton/rackets/defensive?filterBy=&sortBy=typesense.isOOS%3Aasc&page={number}&searchQuery=%2Fbadminton%2Frackets%2Fdefensive%2F&sidebar=false")

    # Get the initial page source
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html.parser")

    # Find elements
    nextpage_elements = driver.find_elements(
        By.CLASS_NAME, "text-sm.font-semibold.leading-snug.text-black.md\\:text-base")

    price_element = driver.find_elements(
        By.CLASS_NAME, "card-price")

    # Store initial elements
    for element in nextpage_elements:
        all_text.append(element.text)

    for element in price_element:
        all_price.append(element.text)

    # Scroll and get more elements
    scroll_down()

    elements_after_scroll = driver.find_elements(
        By.CLASS_NAME, "text-sm.font-semibold.leading-snug.text-black.md\\:text-base")

    price_elements_after_scroll = driver.find_elements(
        By.CLASS_NAME, "card-price")

    # Store elements after scroll
    for element in elements_after_scroll[len(nextpage_elements):]:
        all_text.append(element.text if element.text else "out of stock")

    for element in price_elements_after_scroll[len(price_element):]:
        all_price.append(element.text if element.text else "out of stock")

    number += 1

# Write to Excel
for index, text_content in enumerate(all_text, start=2):
    sheet.cell(row=index, column=1, value=text_content)

for index, price_content in enumerate(all_price, start=2):
    sheet.cell(row=index, column=2, value=price_content)

workbook.save("defensive.xlsx")
driver.quit()
